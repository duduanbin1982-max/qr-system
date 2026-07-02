from io import BytesIO

import openpyxl


def _load_workbook(response):
    return openpyxl.load_workbook(BytesIO(response.data), read_only=True)


def test_users_export_returns_excel_with_chinese_headers(client, auth_headers):
    response = client.get("/api/users/export", headers=auth_headers)

    assert response.status_code == 200, response.get_json(silent=True)
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = _load_workbook(response)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    finally:
        workbook.close()

    assert headers[:4] == ["用户名", "姓名", "员工编号", "电话"]


def test_shipments_export_returns_excel(client, auth_headers):
    response = client.get("/api/shipments/export", headers=auth_headers)

    assert response.status_code == 200, response.get_json(silent=True)
    assert response.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = _load_workbook(response)
    try:
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    finally:
        workbook.close()

    assert headers[:3] == ["出库单号", "客户", "联系人"]
