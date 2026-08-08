import sqlite3
import threading
from concurrent.futures import ThreadPoolExecutor

import pytest
from openpyxl import load_workbook

from factories import create_inventory_item, create_order, ensure_process
from modules.db import get_db
from modules.domain.errors import ConflictError
from modules.services.quality_management_service import QualityManagementService
from modules.services.shipment_service import ShipmentService


CURRENT_USER = {"id": 1, "name": "测试管理员"}


def _shipment_context(db, quantity=10):
    process_id = ensure_process(db, "发货测试工序")
    order_id = create_order(db, [process_id], quantity=quantity, product_code="SHIP-001")
    order = db.execute("SELECT order_no FROM orders WHERE id = ?", (order_id,)).fetchone()
    inventory_id = create_inventory_item(
        db,
        quantity=quantity,
        order_id=order_id,
        product_model="SHIP-001",
        product_name="发货测试产品",
    )
    return order_id, order["order_no"], inventory_id


def _shipment_payload(order_id, order_no, inventory_id, quantity=4, **overrides):
    payload = {
        "customer": "测试客户",
        "order_id": order_id,
        "order_no": order_no,
        "receivable_amount": 100,
        "items": [{
            "inventory_id": inventory_id,
            "product_model": "SHIP-001",
            "product_name": "发货测试产品",
            "quantity": quantity,
            "unit": "件",
            "order_id": order_id,
            "order_no": order_no,
        }],
    }
    payload.update(overrides)
    return payload


def _pass_outgoing_inspection(shipment_id):
    db = get_db()
    row = db.execute(
        "SELECT id FROM quality_inspection_tasks WHERE shipment_id=? AND inspection_type='outgoing'",
        (shipment_id,),
    ).fetchone()
    assert row, "outgoing quality task was not generated"
    task = QualityManagementService.get_task(row["id"])
    measurements = [
        {"item_id": item["id"], "item_code": item["item_code"], "value": item["weight"]}
        for item in task.get("standard_items", [])
    ]
    QualityManagementService.inspect_task(row["id"], {
        "quantity_checked": task["sample_qty"], "quantity_failed": 0,
        "result": "pass", "measurements": measurements,
    }, CURRENT_USER)


def test_create_and_get_shipment_preserves_product_and_order_identity(client):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db)
        payload = _shipment_payload(order_id, order_no, inventory_id)
        payload["order_id"] = order_id + 999
        payload["items"][0].update({
            "order_id": order_id + 999,
            "order_no": "CLIENT-WRONG",
            "product_model": "CLIENT-WRONG",
            "product_name": "客户端错误名称",
        })
        shipment_id, shipment_no = ShipmentService.create_shipment(
            payload, created_by=CURRENT_USER
        )

        shipment = ShipmentService.get_shipment(shipment_id)
        assert shipment["shipment_no"] == shipment_no
        assert shipment["status"] == "pending"
        assert shipment["items"][0]["product_code"] == "SHIP-001"
        assert shipment["items"][0]["order_no"] == order_no
        assert shipment["items"][0]["order_id"] == order_id
        assert shipment["items"][0]["product_name"] == "发货测试产品"
        assert shipment["created_by_id"] == CURRENT_USER["id"]
        assert db.execute(
            "SELECT quantity FROM inventory WHERE id = ?", (inventory_id,)
        ).fetchone()["quantity"] == 10


def test_create_shipment_rejects_empty_items_and_insufficient_stock(client):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db, quantity=2)
        with pytest.raises(ValueError, match="请添加出库产品"):
            ShipmentService.create_shipment({"items": []}, created_by="测试管理员")
        with pytest.raises(ValueError, match="库存不足"):
            ShipmentService.create_shipment(
                _shipment_payload(order_id, order_no, inventory_id, quantity=3),
                created_by="测试管理员",
            )


def test_complete_shipment_deducts_stock_and_updates_delivery_status(client):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db, quantity=4)
        shipment_id, shipment_no = ShipmentService.create_shipment(
            _shipment_payload(order_id, order_no, inventory_id, quantity=4),
            created_by="测试管理员",
        )

        with pytest.raises(ConflictError, match="出库检验任务"):
            ShipmentService.complete_shipment(shipment_id, CURRENT_USER)
        _pass_outgoing_inspection(shipment_id)
        assert ShipmentService.complete_shipment(shipment_id, CURRENT_USER) == shipment_no
        shipment = ShipmentService.get_shipment(shipment_id)
        inventory = db.execute(
            "SELECT quantity FROM inventory WHERE id = ?", (inventory_id,)
        ).fetchone()
        order = db.execute(
            "SELECT delivery_status FROM orders WHERE id = ?", (order_id,)
        ).fetchone()
        assert shipment["status"] == "completed"
        assert inventory["quantity"] == 0
        assert order["delivery_status"] == "全部发货"
        with pytest.raises(ValueError, match="只有待出库单"):
            ShipmentService.complete_shipment(shipment_id, CURRENT_USER)


def test_logistics_receive_and_payment_follow_status_rules(client):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(order_id, order_no, inventory_id), created_by="测试管理员"
        )
        with pytest.raises(ValueError, match="只有已出库单"):
            ShipmentService.receive_shipment(shipment_id, CURRENT_USER)
        with pytest.raises(ValueError, match="只有已出库或已签收"):
            ShipmentService.record_payment(
                shipment_id, CURRENT_USER, 10, idempotency_key="pending-receipt"
            )

        ShipmentService.update_logistics(
            shipment_id, {"logistics_company": "顺丰", "tracking_no": "SF001"}
        )
        _pass_outgoing_inspection(shipment_id)
        ShipmentService.complete_shipment(shipment_id, CURRENT_USER)
        ShipmentService.record_payment(
            shipment_id, CURRENT_USER, 40, "bank", "首款",
            payment_date="2026-07-19", idempotency_key="receipt-1",
        )
        ShipmentService.receive_shipment(shipment_id, CURRENT_USER, "张三", "2026-07-20")
        ShipmentService.record_payment(
            shipment_id, CURRENT_USER, 60, "bank", "尾款",
            payment_date="2026-07-20", idempotency_key="receipt-2",
        )

        shipment = ShipmentService.get_shipment(shipment_id)
        assert shipment["status"] == "received"
        assert shipment["tracking_no"] == "SF001"
        assert shipment["paid_amount"] == 100
        assert shipment["payment_status"] == "paid"
        with pytest.raises(ValueError, match="超出应收"):
            ShipmentService.record_payment(
                shipment_id, CURRENT_USER, 1, idempotency_key="receipt-over"
            )


@pytest.mark.parametrize("action", ["cancel", "delete"])
def test_cancel_or_delete_completed_shipment_restores_stock(client, action):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db, quantity=5)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(order_id, order_no, inventory_id, quantity=5),
            created_by="测试管理员",
        )
        _pass_outgoing_inspection(shipment_id)
        ShipmentService.complete_shipment(shipment_id, CURRENT_USER)

        if action == "cancel":
            ShipmentService.cancel_shipment(shipment_id, CURRENT_USER, "客户撤回")
        else:
            ShipmentService.delete_shipment(shipment_id, CURRENT_USER)
        assert ShipmentService.get_shipment(shipment_id)["status"] == "reversed"
        stock = db.execute(
            "SELECT quantity FROM inventory WHERE id = ?", (inventory_id,)
        ).fetchone()["quantity"]
        assert stock == 5
        assert db.execute(
            "SELECT delivery_status FROM orders WHERE id=?", (order_id,)
        ).fetchone()["delivery_status"] == "pending"


@pytest.mark.parametrize("action", ["cancel", "delete"])
def test_cancel_or_delete_pending_reserved_shipment_releases_reservation(client, action):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db, quantity=5)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(
                order_id, order_no, inventory_id, quantity=4,
                deduction_mode="on_create",
            ),
            created_by="测试管理员",
        )
        reserved = db.execute(
            "SELECT quantity, reserved FROM inventory WHERE id=?", (inventory_id,)
        ).fetchone()
        assert tuple(reserved) == (5, 4)

        if action == "cancel":
            ShipmentService.cancel_shipment(shipment_id, CURRENT_USER, "订单取消")
        else:
            ShipmentService.delete_shipment(shipment_id, CURRENT_USER)

        released = db.execute(
            "SELECT quantity, reserved FROM inventory WHERE id=?", (inventory_id,)
        ).fetchone()
        assert tuple(released) == (5, 0)
        assert [
            row["type"] for row in db.execute(
                "SELECT type FROM inventory_logs WHERE inventory_id=? "
                "AND type IN ('reserve', 'release') ORDER BY id",
                (inventory_id,),
            ).fetchall()
        ] == ["reserve", "release"]


def test_complete_reserved_shipment_consumes_quantity_and_reservation_together(client):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db, quantity=5)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(
                order_id, order_no, inventory_id, quantity=4,
                deduction_mode="on_create",
            ),
            created_by="测试管理员",
        )
        _pass_outgoing_inspection(shipment_id)
        ShipmentService.complete_shipment(shipment_id, CURRENT_USER)

        item = db.execute(
            "SELECT quantity, reserved FROM inventory WHERE id=?", (inventory_id,)
        ).fetchone()
        assert tuple(item) == (1, 0)
        movement = db.execute(
            "SELECT qty_delta, balance_before, balance_after FROM inventory_logs "
            "WHERE idempotency_key LIKE ?",
            ("shipment:%:out",),
        ).fetchone()
        assert tuple(movement) == (-4, 5, 1)


def test_reversed_shipment_cannot_be_completed_again_or_deleted(client):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db, quantity=5)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(order_id, order_no, inventory_id, quantity=5),
            created_by=CURRENT_USER,
        )
        _pass_outgoing_inspection(shipment_id)
        ShipmentService.complete_shipment(shipment_id, CURRENT_USER)
        ShipmentService.cancel_shipment(shipment_id, CURRENT_USER, "回归测试冲销")

        with pytest.raises(ConflictError, match="只有待出库单"):
            ShipmentService.complete_shipment(shipment_id, CURRENT_USER)
        with pytest.raises(ConflictError, match="终态"):
            ShipmentService.delete_shipment(shipment_id, CURRENT_USER)
        assert db.execute(
            "SELECT quantity FROM inventory WHERE id=?", (inventory_id,)
        ).fetchone()["quantity"] == 5
        assert db.execute(
            "SELECT COUNT(*) FROM shipments WHERE id=?", (shipment_id,)
        ).fetchone()[0] == 1


def test_payment_ledger_is_idempotent_and_refund_updates_projection(client):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(order_id, order_no, inventory_id), created_by=CURRENT_USER
        )
        _pass_outgoing_inspection(shipment_id)
        ShipmentService.complete_shipment(shipment_id, CURRENT_USER)

        ShipmentService.record_payment(
            shipment_id, CURRENT_USER, "40.005", payment_date="2026-08-01",
            idempotency_key="payment-idempotent",
        )
        ShipmentService.record_payment(
            shipment_id, CURRENT_USER, "40.005", payment_date="2026-08-01",
            idempotency_key="payment-idempotent",
        )
        ShipmentService.refund_payment(
            shipment_id, CURRENT_USER, "10.01", payment_date="2026-08-02",
            idempotency_key="refund-idempotent",
        )

        shipment = ShipmentService.get_shipment(shipment_id)
        assert shipment["paid_amount"] == 30
        assert shipment["payment_status"] == "partial"
        assert len(shipment["payments"]) == 2
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            db.execute("UPDATE shipment_payments SET remark='changed' WHERE shipment_id=?", (shipment_id,))
        db.rollback()
        with pytest.raises(sqlite3.IntegrityError, match="projection"):
            db.execute("UPDATE shipments SET paid_amount=99 WHERE id=?", (shipment_id,))
        db.rollback()


def test_paid_shipment_requires_refund_before_operational_reversal(client):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(order_id, order_no, inventory_id), created_by=CURRENT_USER
        )
        _pass_outgoing_inspection(shipment_id)
        ShipmentService.complete_shipment(shipment_id, CURRENT_USER)
        ShipmentService.record_payment(
            shipment_id, CURRENT_USER, 20, idempotency_key="paid-before-reverse"
        )

        with pytest.raises(ConflictError, match="请先退款"):
            ShipmentService.cancel_shipment(shipment_id, CURRENT_USER, "客户退货")
        ShipmentService.refund_payment(
            shipment_id, CURRENT_USER, 20, idempotency_key="refund-before-reverse"
        )
        ShipmentService.cancel_shipment(shipment_id, CURRENT_USER, "客户退货")
        assert ShipmentService.get_shipment(shipment_id)["status"] == "reversed"


def test_export_is_not_limited_by_page_size(client):
    with client.application.app_context():
        db = get_db()
        db.executemany(
            "INSERT INTO shipments (shipment_no,customer,status,total_quantity) VALUES (?,?,?,?)",
            [(f"EXPORT-{index:04d}", "导出客户", "pending", 1) for index in range(501)],
        )
        db.commit()
        workbook = load_workbook(ShipmentService.export_shipments())
        assert workbook.active.max_row == 502


def test_unlinked_inventory_requires_explicit_permission(client):
    with client.application.app_context():
        db = get_db()
        inventory_id = create_inventory_item(
            db, quantity=3, order_id=None, product_model="UNLINKED-001",
            product_name="无订单库存",
        )
        payload = _shipment_payload(None, "", inventory_id, quantity=1)
        with pytest.raises(ConflictError, match="无订单发货"):
            ShipmentService.create_shipment(payload, created_by=CURRENT_USER)

        shipment_id, _ = ShipmentService.create_shipment(
            payload, created_by=CURRENT_USER, allow_unlinked=True
        )
        item = ShipmentService.get_shipment(shipment_id)["items"][0]
        assert item["order_id"] is None
        assert item["product_code"] == "UNLINKED-001"


def test_shipment_events_are_immutable(client):
    with client.application.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(order_id, order_no, inventory_id), created_by=CURRENT_USER
        )
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            db.execute(
                "UPDATE shipment_events SET operator_name='changed' WHERE shipment_id=?",
                (shipment_id,),
            )
        db.rollback()
        with pytest.raises(sqlite3.IntegrityError, match="immutable"):
            db.execute("DELETE FROM shipment_events WHERE shipment_id=?", (shipment_id,))
        db.rollback()


def test_concurrent_complete_and_cancel_cannot_corrupt_inventory(client):
    app = client.application
    with app.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db, quantity=5)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(order_id, order_no, inventory_id, quantity=5),
            created_by=CURRENT_USER,
        )
        _pass_outgoing_inspection(shipment_id)

    barrier = threading.Barrier(2)

    def run(action):
        with app.app_context():
            barrier.wait()
            try:
                if action == "complete":
                    ShipmentService.complete_shipment(shipment_id, CURRENT_USER)
                else:
                    ShipmentService.cancel_shipment(
                        shipment_id, CURRENT_USER, "并发取消测试"
                    )
                return "success"
            except ValueError:
                return "conflict"

    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(run, ("complete", "cancel")))

    with app.app_context():
        db = get_db()
        shipment = ShipmentService.get_shipment(shipment_id)
        assert shipment["status"] in ("cancelled", "reversed")
        assert db.execute(
            "SELECT quantity FROM inventory WHERE id=?", (inventory_id,)
        ).fetchone()["quantity"] == 5
        assert "success" in results


def test_concurrent_receipts_serialize_against_receivable(client):
    app = client.application
    with app.app_context():
        db = get_db()
        order_id, order_no, inventory_id = _shipment_context(db)
        shipment_id, _ = ShipmentService.create_shipment(
            _shipment_payload(order_id, order_no, inventory_id), created_by=CURRENT_USER
        )
        _pass_outgoing_inspection(shipment_id)
        ShipmentService.complete_shipment(shipment_id, CURRENT_USER)

    barrier = threading.Barrier(2)

    def receive(key):
        with app.app_context():
            barrier.wait()
            try:
                ShipmentService.record_payment(
                    shipment_id, CURRENT_USER, 60, idempotency_key=key
                )
                return "success"
            except ConflictError:
                return "conflict"

    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(receive, ("concurrent-receipt-1", "concurrent-receipt-2")))

    with app.app_context():
        shipment = ShipmentService.get_shipment(shipment_id)
        assert sorted(results) == ["conflict", "success"]
        assert shipment["paid_amount"] == 60
        assert shipment["payment_status"] == "partial"
        assert len(shipment["payments"]) == 1
