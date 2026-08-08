"""
qr-system — 库存管理 Service 层

从 routes/inventory.py 提取全部业务逻辑。
"""
from modules.domain.errors import ConflictError, NotFoundError
from datetime import datetime
from modules.services import BaseService
from modules.repositories.inventory_repository import InventoryRepository
from modules.services.inventory_posting_service import InventoryPostingService


class InventoryService:
    """库存管理业务逻辑。"""

    @staticmethod
    def list_items(keyword='', low_stock=False, location='', page=1, limit=100):
        """库存列表（搜索 + 低库存筛选 + 分页）。"""
        where, params = InventoryRepository.build_item_filters(keyword, low_stock, location)
        total = InventoryRepository.count_items(where, params)
        rows, size = InventoryRepository.list_items_paginated(where, params, page, limit)
        return {'items': [dict(r) for r in rows], 'total': total, 'page': page, 'limit': size}

    @staticmethod
    def create_item(data):
        """Create a zero-balance item, then post any opening balance."""
        model = (data.get('product_model') or '').strip()
        if not model:
            raise ValueError('产品型号不能为空')
        try:
            opening_quantity = float(data.get('quantity', 0) or 0)
        except (TypeError, ValueError):
            raise ValueError('初始数量必须为数字')
        if opening_quantity < 0:
            raise ValueError('初始数量不能为负数')
        order_id = data.get('order_id') or None
        with BaseService.transaction() as txn:
            if InventoryRepository.find_duplicate_model_txn(model, order_id, 0, txn):
                raise ConflictError('产品型号已存在')
            item_id = InventoryRepository.insert_txn(
                model,
                data.get('product_name', ''),
                data.get('specification', ''),
                data.get('safe_stock', 0),
                data.get('location', ''),
                data.get('unit', '件'),
                data.get('remark', ''),
                data.get('category', ''),
                data.get('unit_cost', 0),
                order_id,
                txn,
            )
            if opening_quantity:
                InventoryPostingService.post(
                    item_id,
                    opening_quantity,
                    'opening_balance',
                    remark='创建库存时录入期初余额',
                    source_type='inventory_create',
                    source_id=item_id,
                    idempotency_key='inventory:%s:opening' % item_id,
                    db=txn,
                )
            return item_id

    @staticmethod
    def update_item(item_id, data):
        """Update item metadata without mutating its audited balance."""
        current = InventoryRepository.find_item_by_id(item_id)
        if not current:
            raise NotFoundError('库存不存在')
        model = (data.get('product_model') or '').strip()
        if not model:
            raise ValueError('产品型号不能为空')
        with BaseService.transaction() as txn:
            current = InventoryRepository.find_item_by_id(item_id, db=txn)
            if not current:
                raise NotFoundError('库存不存在')
            if InventoryRepository.find_duplicate_model_txn(
                model, current['order_id'], item_id, txn
            ):
                raise ConflictError('该订单下产品型号已存在')
            cursor = InventoryRepository.update_item_txn(
                item_id,
                model,
                data.get('product_name', ''),
                data.get('specification', ''),
                data.get('safe_stock', 0),
                data.get('location', ''),
                data.get('unit', '件'),
                data.get('remark', ''),
                data.get('category', ''),
                data.get('unit_cost', 0),
                txn,
            )
            if cursor.rowcount != 1:
                raise ConflictError('库存状态已变化，请刷新后重试')

    @staticmethod
    def delete_item(item_id):
        """Archive a zero-balance item while retaining its ledger."""
        item = InventoryRepository.find_item_for_delete(item_id)
        if not item or item['deleted_at']:
            raise NotFoundError('库存不存在')
        if float(item['quantity'] or 0) != 0:
            raise ConflictError('库存不为零，不能停用')
        if float(item['reserved'] or 0) != 0:
            raise ConflictError('仍有预留库存，不能停用')
        with BaseService.transaction() as txn:
            cursor = InventoryRepository.archive_item_txn(item_id, txn)
            if cursor.rowcount != 1:
                raise ConflictError('库存状态已变化，请刷新后重试')

    @staticmethod
    def stock_in(inv_id, qty, order_id=None, order_no='', remark='',
                 operator_id=None, operator_name='', lot_no='', serial_no='',
                 idempotency_key=''):
        """入库操作。"""
        try:
            qty = float(qty)
        except (TypeError, ValueError):
            raise ValueError('数量必须为数字')
        if qty <= 0:
            raise ValueError('参数错误')
        return InventoryPostingService.post(
            inv_id, qty, 'in', order_id=order_id, order_no=order_no,
            remark=remark, operator_id=operator_id, operator_name=operator_name,
            lot_no=lot_no, serial_no=serial_no, source_type='manual',
            idempotency_key=idempotency_key,
        )

    @staticmethod
    def stock_out(inv_id, qty, order_id=None, order_no='', remark='',
                  operator_id=None, operator_name='', lot_no='', serial_no='',
                  idempotency_key=''):
        """出库操作（原子扣减 + 防超卖）。"""
        try:
            qty = float(qty)
        except (TypeError, ValueError):
            raise ValueError('数量必须为数字')
        if qty <= 0:
            raise ValueError('参数错误')
        return InventoryPostingService.post(
            inv_id, -qty, 'out', order_id=order_id, order_no=order_no,
            remark=remark, operator_id=operator_id, operator_name=operator_name,
            lot_no=lot_no, serial_no=serial_no, source_type='manual',
            idempotency_key=idempotency_key,
        )

    @staticmethod
    def get_logs(inv_id='', type_filter='', page=1, limit=20, date_from='', date_to=''):
        """库存流水（分页）。"""
        total = InventoryRepository.count_logs(inv_id, type_filter, date_from, date_to)
        rows = InventoryRepository.list_logs(inv_id, type_filter, date_from, date_to, page, limit)
        return {'logs': [dict(r) for r in rows], 'total': total}

    @staticmethod
    def get_alerts():
        """库存预警列表。"""
        rows = InventoryRepository.list_alerts()
        return {'alerts': [dict(r) for r in rows]}

    @staticmethod
    def stock_adjust(inv_id, actual_qty, operator_id=None, operator_name='', remark=''):
        try:
            actual_qty = float(actual_qty)
        except (TypeError, ValueError):
            raise ValueError('盘点数量必须为数字')
        if actual_qty < 0:
            raise ValueError('盘点数量不能为负数')
        with BaseService.transaction() as txn:
            inv = InventoryRepository.find_adjustment_item(inv_id, db=txn)
            if not inv:
                raise NotFoundError('库存记录不存在')
            current = float(inv['quantity'] or 0)
            diff = actual_qty - current
            if diff == 0:
                return {'adjusted': False, 'message': '库存数量一致，无需调整'}
            InventoryPostingService.post(
                inv_id, diff, 'count_gain' if diff > 0 else 'count_loss',
                remark='盘点调整: %s (原%s→现%s, 差额%+g)' % (
                    remark or '系统调整', current, actual_qty, diff
                ),
                operator_id=operator_id,
                operator_name=operator_name,
                source_type='manual_count',
                db=txn,
            )
        return {'adjusted': True, 'product_model': inv['product_model'],
                'old_qty': current, 'new_qty': actual_qty, 'diff': diff}


    # P2: ABC
    @staticmethod
    def classify_abc():
        rows = InventoryRepository.list_abc_rows()
        if not rows:
            return {"message": "无库存数据"}
        total_value = sum(r["out_value"] for r in rows)
        total = len(rows)
        if total_value == 0:
            a_cut = max(1, int(total * 0.2))
            b_cut = max(a_cut + 1, int(total * 0.5))
            with BaseService.transaction() as txn:
                for i, r in enumerate(rows):
                    cat = "A" if i < a_cut else ("B" if i < b_cut else "C")
                    InventoryRepository.update_category_txn(r["id"], cat, txn)
            return {"classified": total, "a_count": a_cut, "b_count": b_cut - a_cut, "c_count": total - b_cut}
        cum = 0
        a_cut = b_cut = 0
        with BaseService.transaction() as txn:
            for i, r in enumerate(rows):
                cum += r["out_value"]
                pct = cum / total_value
                cat = "A" if pct <= 0.8 else ("B" if pct <= 0.95 else "C")
                if cat == "A" and a_cut == 0: a_cut = i + 1
                if cat == "B" and b_cut == 0: b_cut = i + 1
                InventoryRepository.update_category_txn(r["id"], cat, txn)
        if b_cut == 0: b_cut = a_cut + 1
        return {"classified": total, "a_count": a_cut, "b_count": b_cut - a_cut, "c_count": total - b_cut + 1}

    @staticmethod
    def get_turnover():
        rows = InventoryRepository.list_turnover_rows()
        result = []
        for r in rows:
            turnover = round(r["total_out"] / max(r["current_stock"], 1), 2)
            result.append({"id": r["id"], "product_model": r["product_model"], "product_name": r["product_name"], "current_stock": r["current_stock"], "total_out": r["total_out"], "total_in": r["total_in"], "unit_cost": r["unit_cost"], "turnover_rate": turnover, "status": "高周转" if turnover > 3 else ("正常" if turnover > 1 else "低周转")})
        return {"items": result}

    @staticmethod
    def suggest_safe_stock():
        rows = InventoryRepository.list_safe_stock_suggestion_rows()
        suggestions = []
        for r in rows:
            daily_avg = round(r["month_out"] / 30, 1)
            suggested = max(1, int(daily_avg * 7))
            suggestions.append({"id": r["id"], "product_model": r["product_model"], "product_name": r["product_name"], "current_safe_stock": r["current_safe"], "suggested_safe_stock": suggested, "daily_avg_consumption": daily_avg, "current_quantity": r["quantity"], "need_adjust": abs(suggested - r["current_safe"]) > 0})
        return suggestions

    @staticmethod
    def get_batch_tracking(item_id=None, lot_no=None):
        movements = InventoryRepository.list_batch_movements(item_id=item_id)
        batches_by_inventory = {}
        batches = []
        for row in movements:
            movement = dict(row)
            inventory_id = movement["inventory_id"]
            delta = float(movement.get("qty_delta") or 0)
            movement_lot = movement.get("lot_no") or ""
            inventory_batches = batches_by_inventory.setdefault(inventory_id, [])
            if delta > 0:
                batch = {
                    **movement,
                    "received_quantity": delta,
                    "remaining": delta,
                    "related_outs": [],
                }
                inventory_batches.append(batch)
                batches.append(batch)
                continue
            if delta >= 0:
                continue

            remaining_out = abs(delta)
            candidates = [
                batch for batch in inventory_batches
                if batch["remaining"] > 0
                and (not movement_lot or (batch.get("lot_no") or "") == movement_lot)
            ]
            for batch in candidates:
                allocated = min(batch["remaining"], remaining_out)
                if allocated <= 0:
                    continue
                batch["remaining"] -= allocated
                batch["related_outs"].append({
                    **movement,
                    "allocated_quantity": allocated,
                })
                remaining_out -= allocated
                if remaining_out <= 0:
                    break

        if lot_no:
            batches = [batch for batch in batches if (batch.get("lot_no") or "") == lot_no]
        batches.reverse()
        return {"batches": batches[:100]}

    @staticmethod
    def get_locations():
        rows = InventoryRepository.list_locations()
        return {"locations": [dict(r) for r in rows]}

    @staticmethod
    def update_location(item_ids, new_location):
        if not item_ids or not new_location:
            raise ValueError("请提供物料ID和目标库位")
        with BaseService.transaction() as txn:
            for iid in item_ids:
                InventoryRepository.update_location_txn(iid, new_location, txn)
        return {"updated": len(item_ids), "location": new_location}

    @staticmethod
    def create_count_task(user_id=None, user_name=''):
        with BaseService.transaction() as txn:
            open_task = InventoryRepository.find_open_count_task(db=txn)
            if open_task:
                raise ConflictError('已有未完成盘点任务 %s' % open_task['task_no'])
            if InventoryRepository.count_inventory_items(db=txn) == 0:
                raise ValueError('没有可盘点的库存')
            task_no = 'CT-' + datetime.now().strftime('%Y%m%d%H%M%S%f')
            task_id = InventoryRepository.create_count_task_txn(
                task_no, user_id, user_name, txn
            )
        return InventoryService.get_count_status(task_id)

    @staticmethod
    def get_count_status(task_id=None):
        task = (
            InventoryRepository.find_count_task(task_id)
            if task_id else InventoryRepository.find_latest_count_task()
        )
        if not task:
            return {"task": None, "items": [], "total": 0, "done": 0,
                    "pending": 0, "progress_pct": 0}
        items = [dict(row) for row in InventoryRepository.list_count_task_items(task['id'])]
        done = sum(1 for item in items if item['status'] != 'pending')
        total = len(items)
        return {
            "task": dict(task),
            "items": items,
            "total": total,
            "done": done,
            "pending": total - done,
            "progress_pct": round(done / max(total, 1) * 100, 1),
        }

    @staticmethod
    def export_inventory(keyword='', low_stock=False):
        from modules.export_utils import style_header, auto_width, THIN_BORDER, CELL_ALIGN
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from io import BytesIO

        result = InventoryService.list_items(keyword=keyword, low_stock=low_stock, page=1, limit=99999)
        items = result.get('items', [])

        wb = Workbook()
        ws = wb.active
        ws.title = '库存清单'

        headers = ['产品名称', '订单号', '客户', '产品型号', '规格', '当前库存', '安全库存', '状态', '存放位置', '单位', '备注', '更新时间']
        style_header(ws, headers)

        for row_idx, item in enumerate(items, 2):
            status = '⚠低库存' if item.get('is_low') else '正常'
            vals = [
                item.get('product_name', ''), item.get('order_no', ''),
                item.get('customer', ''), item.get('product_model', ''),
                item.get('specification', ''), item.get('quantity', 0),
                item.get('safe_stock', 0), status,
                item.get('location', ''), item.get('unit', ''),
                item.get('remark', ''), (item.get('updated_at') or '')[:19]
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGN
                if status == '⚠低库存':
                    cell.font = Font(name='Microsoft YaHei', color='FF0000')

        auto_width(ws)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_logs(inv_id='', type_filter='', date_from='', date_to=''):
        from modules.export_utils import style_header, auto_width, THIN_BORDER, CELL_ALIGN
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from io import BytesIO

        result = InventoryService.get_logs(inv_id=inv_id, type_filter=type_filter,
                                           date_from=date_from, date_to=date_to, page=1, limit=99999)
        items = result.get('logs', [])

        wb = Workbook()
        ws = wb.active
        ws.title = '库存流水'

        headers = ['时间', '类型', '产品型号', '产品名称', '数量', '订单号', '操作人', '备注']
        style_header(ws, headers)

        type_map = {'in': '入库', 'out': '出库', 'adjust': '盘点调整'}
        for row_idx, item in enumerate(items, 2):
            vals = [
                (item.get('created_at') or '')[:19],
                type_map.get(item.get('type', ''), item.get('type', '')),
                item.get('product_model', ''), item.get('product_name', ''),
                item.get('qty_delta', 0), item.get('order_no', ''),
                item.get('operator_name', ''), item.get('remark', '')
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = CELL_ALIGN

        auto_width(ws)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def submit_count(task_id, item_id, actual_qty, remark="", user_id=None, user_name=""):
        try:
            actual_qty = float(actual_qty)
        except (TypeError, ValueError):
            raise ValueError("盘点数量必须为数字")
        if actual_qty < 0:
            raise ValueError("盘点数量不能为负数")
        with BaseService.transaction() as txn:
            task = InventoryRepository.find_count_task(task_id, db=txn)
            if not task:
                raise NotFoundError("盘点任务不存在")
            if task["status"] != "counting":
                raise ConflictError("盘点任务已提交，不能继续录入")
            count_items = InventoryRepository.list_count_task_items(task_id, db=txn)
            count_item = next(
                (row for row in count_items if row["inventory_id"] == item_id), None
            )
            if not count_item:
                raise NotFoundError("盘点明细不存在")
            cursor = InventoryRepository.update_count_item_txn(
                task_id, item_id, actual_qty, remark, user_id, user_name, txn
            )
            if cursor.rowcount != 1:
                raise ConflictError("盘点明细状态已变化，请刷新后重试")
            diff = actual_qty - float(count_item["book_quantity"] or 0)
        return {"ok": True, "old_qty": count_item["book_quantity"],
                "new_qty": actual_qty, "diff": diff}

    @staticmethod
    def approve_count_task(task_id, user_id=None, user_name=""):
        with BaseService.transaction() as txn:
            task = InventoryRepository.find_count_task(task_id, db=txn)
            if not task:
                raise NotFoundError("盘点任务不存在")
            if task["status"] != "submitted":
                raise ConflictError("盘点尚未全部录入或已经审批")
            items = InventoryRepository.list_count_task_items(task_id, db=txn)
            for item in items:
                if item["actual_quantity"] is None:
                    raise ConflictError("盘点任务仍有未录入明细")
                current = InventoryRepository.find_item_by_id(item["inventory_id"], db=txn)
                if not current:
                    raise ConflictError("盘点库存已停用，请重新创建任务")
                if abs(
                    float(current["quantity"] or 0) - float(item["book_quantity"] or 0)
                ) > 0.0000001:
                    raise ConflictError(
                        "%s 在盘点期间发生库存变动，请重新盘点"
                        % item["product_model"]
                    )
                diff = float(item["difference"] or 0)
                movement_id = None
                if diff:
                    movement = InventoryPostingService.post(
                        item["inventory_id"],
                        diff,
                        "count_gain" if diff > 0 else "count_loss",
                        remark="盘点任务 %s 审批过账%s" % (
                            task["task_no"],
                            ("：" + item["remark"]) if item["remark"] else "",
                        ),
                        operator_id=user_id,
                        operator_name=user_name,
                        source_type="count_task",
                        source_id=item["id"],
                        idempotency_key="count:%s:item:%s:post" % (task_id, item["id"]),
                        db=txn,
                    )
                    movement_id = movement["id"]
                InventoryRepository.mark_count_item_posted_txn(
                    item["id"], movement_id, txn
                )
                InventoryRepository.mark_inventory_counted_txn(
                    item["inventory_id"], txn
                )
            cursor = InventoryRepository.approve_count_task_txn(
                task_id, user_id, user_name, txn
            )
            if cursor.rowcount != 1:
                raise ConflictError("盘点任务状态已变化，请刷新后重试")
        return InventoryService.get_count_status(task_id)

    @staticmethod
    def get_impact(item_id):
        item = InventoryRepository.find_item_for_delete(item_id)
        if not item:
            raise NotFoundError("item not found")
        log_count = InventoryRepository.count_item_logs(item_id)
        order_count = InventoryRepository.count_linked_orders(item_id)
        warnings = []
        if log_count > 0:
            warnings.append("将保留 %s 条历史流水" % log_count)
        if order_count > 0:
            warnings.append("关联 %s 个订单" % order_count)
        shipment_count = InventoryRepository.count_linked_shipment_items(item_id)
        if shipment_count > 0:
            warnings.append("关联 %s 条出库明细" % shipment_count)
        can_archive = float(item["quantity"] or 0) == 0 and float(item["reserved"] or 0) == 0
        return {
            "item": dict(item),
            "log_count": log_count,
            "order_count": order_count,
            "shipment_count": shipment_count,
            "can_delete": can_archive,
            "can_archive": can_archive,
            "warnings": [w for w in warnings if w]
        }

    @staticmethod
    def get_stats():
        """库存统计（2次查询替代4次）。"""
        today = datetime.now().strftime('%Y-%m-%d')
        inv_stats = InventoryRepository.get_inventory_stats()
        today_stats = InventoryRepository.get_today_stats(today)
        return {
            'total_items': inv_stats['total_items'] or 0,
            'total_quantity': inv_stats['total_quantity'] or 0,
            'low_stock': inv_stats['low_stock'] or 0,
            'today_in': today_stats['today_in'] or 0,
            'today_out': today_stats['today_out'] or 0,
        }
