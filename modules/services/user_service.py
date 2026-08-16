"""qr-system - User Service (Repository-pattern refactor)

All business logic (validation, bcrypt, secrets) stays here.
All SQL delegated to UserRepository.
"""
from modules.domain.errors import ConflictError, NotFoundError, ValidationError
import bcrypt
import secrets
import os
import tempfile
import uuid
import zipfile
from io import BytesIO

from openpyxl import Workbook
from modules.services import BaseService
from modules.repositories.user_repository import UserRepository
from modules.repositories.performance_assignment_repository import (
    PerformanceAssignmentRepository,
)
from modules.services.performance_assignment_service import (
    PerformanceAssignmentService,
)
from modules.services.administrator_policy import AdministratorPolicy
from modules.config import (
    EMPLOYEE_DOCUMENT_ALLOWED_EXTENSIONS,
    EMPLOYEE_DOCUMENT_MAX_BYTES,
)

USER_IMPORT_FIELDS = [
    "username", "name", "employee_no", "phone", "email", "nickname",
    "position_name", "role", "password", "process_names",
]
USER_IMPORT_FIELD_MAP = {field: field for field in USER_IMPORT_FIELDS}
USER_IMPORT_EN_MAP = {field.lower(): field for field in USER_IMPORT_FIELDS}


class UserService:
    """User management business logic. All methods are static."""

    @staticmethod
    def _validate_process_ids(data):
        """Shared process validation - auto-filter invalid process IDs (self-healing)."""
        if "process_ids" in data and data["process_ids"]:
            id_list = [int(x.strip()) for x in data["process_ids"].split(",") if x.strip()]
            if id_list:
                valid_ids = UserRepository.validate_process_ids(id_list)
                filtered = [str(x) for x in id_list if x in valid_ids]
                data["process_ids"] = ",".join(filtered) if filtered else ""
                data["_valid_process_ids"] = [int(x) for x in filtered] if filtered else []

    @staticmethod
    def _validate_username(username, name):
        if not username or not name:
            raise ValueError("Username and name cannot be empty")
        if len(username) < 2 or len(username) > 32:
            raise ValueError("Username must be 2-32 characters")
        if not all(char.isalnum() or char in "_-." for char in username):
            raise ValueError("Username can only contain letters, digits, underscores, hyphens and dots")

    @staticmethod
    def _resolve_role(data, db):
        role_id = data.get("role_id")
        role_code = (data.get("role") or "").strip()
        if role_id:
            role_code = UserRepository.find_role_code_by_id(role_id, db=db)
            if not role_code:
                raise NotFoundError("Specified role does not exist")
            return role_id, role_code
        if not role_code:
            role_code = "worker"
        role_row = UserRepository.find_role_by_code(role_code, db=db)
        if not role_row:
            role_row = UserRepository.find_role_by_code("worker", db=db)
            role_code = "worker"
        return (role_row[0] if role_row else 2), role_code

    @staticmethod
    def _ensure_admin_creator(role_code, caller_id, db):
        if role_code == "worker":
            return
        AdministratorPolicy.require_actual_admin(
            caller_id,
            db,
            message="仅系统管理员可以创建特权账号",
        )

    @staticmethod
    def _next_employee_no(data, db):
        employee_no = (data.get("employee_no") or "").strip()
        if employee_no:
            if UserRepository.find_user_by_employee_no(employee_no, db=db):
                raise ConflictError("Employee number already exists")
            return employee_no
        next_no = UserRepository.get_next_employee_no(db=db)
        employee_no = str(next_no).zfill(4)
        while UserRepository.check_employee_no_exists(employee_no, db=db):
            next_no += 1
            employee_no = str(next_no).zfill(4)
        return employee_no

    @staticmethod
    def _hash_or_generate_password(data):
        raw_pw = (data.get("password") or "").strip() or secrets.token_urlsafe(8)
        if data.get("password") and len(raw_pw) < 6:
            raise ValueError("Password must be at least 6 characters")
        return raw_pw, bcrypt.hashpw(raw_pw.encode(), bcrypt.gensalt(rounds=12)).decode()

    @staticmethod
    def _resolve_update_role(uid, data, old_role, current_user_id, db):
        new_role_id = None
        new_role_code = None
        if ("role" in data and data["role"] != old_role) or ("role_id" in data):
            new_role_id = data.get("role_id")
            if new_role_id:
                new_role_code = UserRepository.find_role_code_by_id(new_role_id, db=db)
                if not new_role_code:
                    raise NotFoundError("Specified role does not exist")
            else:
                new_role_code = (data.get("role") or "worker").strip() or "worker"
                new_role_id = UserRepository.find_role_id_by_code(new_role_code, db=db)
                if not new_role_id:
                    raise NotFoundError("Specified role does not exist")
            group_row = UserRepository.get_role_group_name(new_role_id, db=db)
            if group_row:
                data["group_name"] = group_row[0]
            data["role"] = new_role_code

        new_role_id_for_check = new_role_id
        if not new_role_id_for_check and "role" in data:
            new_role_id_for_check = UserRepository.find_role_id_by_code(data["role"], db=db)
        old_role_id_for_check = UserRepository.find_role_id_by_code(old_role, db=db)
        if new_role_id_for_check is not None and new_role_id_for_check != old_role_id_for_check:
            if current_user_id is None:
                raise ValueError("Only administrators can change roles")
            if current_user_id == uid:
                raise ConflictError("Cannot change your own role")
            if not UserRepository.check_admin_role(current_user_id, db=db):
                raise ValueError("Only administrators can change roles")
        return new_role_id, new_role_code

    @staticmethod
    def _build_update_fields(data):
        if "employee_no" in data and not data["employee_no"]:
            del data["employee_no"]
        sets = []
        params = []
        for field in ["name", "nickname", "email", "group_name", "role", "employee_no",
                      "marker", "phone", "status", "position_id", "department_id"]:
            if field in data:
                sets.append(field + " = ?")
                params.append(data[field])
        if "password" in data and data["password"]:
            sets.append("password = ?")
            params.append(bcrypt.hashpw(data["password"].encode(), bcrypt.gensalt(rounds=12)).decode())
        if not sets:
            raise ValueError("No update fields provided")
        return sets, params

    @staticmethod
    def _sync_user_processes(uid, data, txn):
        if "process_ids" not in data and "_valid_process_ids" not in data:
            return
        UserRepository.delete_user_processes_txn(uid, db=txn)
        valid_pids = data.get("_valid_process_ids", [])
        if not valid_pids and data.get("process_ids"):
            valid_pids = [int(x.strip()) for x in data["process_ids"].split(",") if x.strip()]
        for pid in valid_pids:
            UserRepository.insert_user_process_txn(uid, pid, db=txn)

    @staticmethod
    def _apply_role_update(uid, old_role, new_role_id, new_role_code, txn):
        if new_role_id is None:
            return
        if new_role_code != "admin":
            remaining = UserRepository.count_admin_roles_excluding(uid, db=txn)
            if remaining == 0:
                remaining_users = UserRepository.count_admin_users_excluding(uid, db=txn)
                if remaining_users == 0:
                    raise ConflictError("Cannot remove the last administrator")
        old_role_id = UserRepository.find_role_id_by_code(old_role, db=txn) or 2
        UserRepository.delete_user_role_txn(uid, old_role_id, db=txn)
        UserRepository.insert_user_role_txn(uid, new_role_id, db=txn)

    @staticmethod
    def _audit_user_update(uid, data, existing, current_user_id, db):
        changed = []
        field_labels = {
            "name": "Name", "nickname": "Nickname", "email": "Email", "phone": "Phone",
            "role": "Role", "employee_no": "Employee No", "marker": "Marker", "group_name": "Role Group",
            "position_id": "Position ID", "department_id": "Department ID", "status": "Status"
        }
        for field, label in field_labels.items():
            old_val = existing[field] if field in existing.keys() else None
            if field not in data:
                continue
            new_val = data.get(field)
            if str(old_val) != str(new_val):
                changed.append(label + ": " + str(old_val) + " -> " + str(new_val))
        if "password" in data and data["password"]:
            changed.append("Password: changed")
        if changed:
            UserRepository.insert_audit_log_txn(
                current_user_id, "update_user", "user", uid, "; ".join(changed), db=db
            )

    # ============================================================
    # Query
    # ============================================================

    @staticmethod
    def list_users(page=1, limit=20, role_filter="", role_not="", keyword="", status=""):
        """Paginated user list."""
        return UserRepository.list_users(
            page=page, limit=limit, role_filter=role_filter,
            role_not=role_not, keyword=keyword, status=status
        )

    # ============================================================
    # Role administration
    # ============================================================

    @staticmethod
    def _normalize_ids(values, label):
        if not isinstance(values, list):
            raise ValueError(label + " must be a list")
        try:
            normalized = sorted({int(value) for value in values})
        except (TypeError, ValueError):
            raise ValueError(label + " must contain integers") from None
        if any(value <= 0 for value in normalized):
            raise ValueError(label + " must contain positive integers")
        return normalized

    @staticmethod
    def _require_role_administrator(actor_id, db):
        if not actor_id or not UserRepository.check_admin_role(actor_id, db=db):
            raise PermissionError("users:admin permission is required")

    @staticmethod
    def _prepare_role_change(user_id, role_ids, action, db):
        if action not in {"set", "add", "remove"}:
            raise ValueError("Invalid role action")
        target = UserRepository.find_user_by_id_for_update(user_id, db=db)
        if not target:
            raise NotFoundError("User not found")
        if target["status"] == "deleted":
            raise ConflictError("Cannot change roles for a deleted user")

        requested_rows = UserRepository.find_active_roles_by_ids(role_ids, db=db)
        if len(requested_rows) != len(role_ids):
            valid_ids = {row["id"] for row in requested_rows}
            invalid = [str(role_id) for role_id in role_ids if role_id not in valid_ids]
            raise ValueError("Invalid or inactive role IDs: " + ", ".join(invalid))

        current_rows = UserRepository.get_user_role_rows(user_id, db=db)
        role_codes = {row["id"]: row["code"] for row in current_rows}
        role_codes.update({row["id"]: row["code"] for row in requested_rows})
        current_ids = {row["id"] for row in current_rows}
        requested_ids = set(role_ids)
        if action == "set":
            proposed_ids = requested_ids
        elif action == "add":
            proposed_ids = current_ids | requested_ids
        else:
            proposed_ids = current_ids - requested_ids
        if not proposed_ids:
            raise ConflictError("A user must retain at least one role")

        current_codes = {row["code"] for row in current_rows}
        proposed_codes = {role_codes[role_id] for role_id in proposed_ids}
        if (
            "admin" in current_codes
            and "admin" not in proposed_codes
            and UserRepository.count_admin_roles_excluding(user_id, db=db) == 0
        ):
            raise ConflictError("Cannot remove the last administrator")
        return sorted(current_ids), sorted(proposed_ids), sorted(current_codes), sorted(proposed_codes)

    @staticmethod
    def _apply_role_change(user_id, requested_ids, proposed_ids, proposed_codes, action, actor_id, db):
        if action == "set":
            UserRepository.replace_user_roles_txn(user_id, proposed_ids, db=db)
        elif action == "add":
            UserRepository.add_user_roles_txn(user_id, requested_ids, db=db)
        else:
            UserRepository.remove_user_roles_txn(user_id, requested_ids, db=db)
        UserRepository.update_user_base_role_txn(
            user_id, "admin" if "admin" in proposed_codes else "worker", db=db
        )
        UserRepository.insert_audit_log_txn(
            actor_id,
            "set_user_roles" if action == "set" else "batch_set_roles",
            "user",
            user_id,
            "action=" + action + "; roles=" + ",".join(proposed_codes),
            db=db,
        )

    @staticmethod
    def set_user_roles(user_id, role_ids, actor_id):
        role_ids = UserService._normalize_ids(role_ids, "role_ids")
        if not role_ids:
            raise ConflictError("A user must retain at least one role")
        if user_id == actor_id:
            raise ConflictError("Cannot change your own roles")
        with BaseService.transaction() as txn:
            UserService._require_role_administrator(actor_id, txn)
            _, proposed_ids, _, proposed_codes = UserService._prepare_role_change(
                user_id, role_ids, "set", txn
            )
            UserService._apply_role_change(
                user_id, role_ids, proposed_ids, proposed_codes, "set", actor_id, txn
            )
        return {"user_id": user_id, "role_ids": proposed_ids}

    @staticmethod
    def batch_set_user_roles(user_ids, role_ids, action, actor_id):
        user_ids = UserService._normalize_ids(user_ids, "user_ids")
        role_ids = UserService._normalize_ids(role_ids, "role_ids")
        if not user_ids or not role_ids:
            raise ValueError("user_ids and role_ids are required")
        if actor_id in user_ids:
            raise ConflictError("Cannot change your own roles")
        changed = 0
        with BaseService.transaction() as txn:
            UserService._require_role_administrator(actor_id, txn)
            for user_id in user_ids:
                _, proposed_ids, _, proposed_codes = UserService._prepare_role_change(
                    user_id, role_ids, action, txn
                )
                UserService._apply_role_change(
                    user_id, role_ids, proposed_ids, proposed_codes, action, actor_id, txn
                )
                changed += 1
        return changed

    # ============================================================
    # Create
    # ============================================================

    @staticmethod
    def create_user(data):
        """Create a new user. Returns (uid, password)."""
        username = data.get("username", "").strip()
        name = data.get("name", "").strip()
        status = AdministratorPolicy.normalize_status(data.get("status", "active"))
        if not name and data.get("role") == "admin":
            name = username
        UserService._validate_username(username, name)

        db = BaseService.db()
        role_id, role_code = UserService._resolve_role(data, db)
        UserService._ensure_admin_creator(role_code, data.get("_caller_user_id"), db)

        # Uniqueness check
        if UserRepository.find_user_by_username(username, db=db):
            raise ConflictError("Username already exists")

        # Position validation
        position_id = data.get("position_id")
        if position_id:
            if not UserRepository.find_position_by_id(position_id, db=db):
                raise NotFoundError("Specified position does not exist")
        department_id = data.get("department_id")
        if department_id and not UserRepository.find_department_by_id(department_id, db=db):
            raise NotFoundError("Specified department does not exist")

        # Process validation
        UserService._validate_process_ids(data)

        raw_pw, pw = UserService._hash_or_generate_password(data)

        with BaseService.transaction() as txn:
            UserService._ensure_admin_creator(
                role_code, data.get("_caller_user_id"), txn
            )
            if UserRepository.find_user_by_username(username, db=txn):
                raise ConflictError("Username already exists")
            data["employee_no"] = UserService._next_employee_no(data, txn)
            uid = UserRepository.insert_user_txn(
                username=username, pw_hash=pw, name=name,
                nickname=data.get("nickname", ""),
                email=data.get("email", ""),
                group_name=data.get("group_name", '员工组'),
                role=role_code,
                employee_no=data.get("employee_no", ""),
                marker=(data.get("marker") or "").strip(),
                phone=data.get("phone", ""),
                position_id=position_id or None,
                department_id=department_id or None,
                status=status,
                db=txn
            )
            UserRepository.insert_user_role_txn(uid, role_id, db=txn)
            # Sync process assignments
            valid_pids = data.get("_valid_process_ids", [])
            if not valid_pids and data.get("process_ids"):
                valid_pids = [int(x.strip()) for x in data["process_ids"].split(",") if x.strip()]
            for pid in valid_pids:
                UserRepository.insert_user_process_txn(uid, pid, db=txn)
            PerformanceAssignmentService.record_initial_assignment(
                uid,
                created_by=data.get("_caller_user_id"),
                source_type="user_created",
                db=txn,
            )
            UserRepository.insert_audit_log_txn(
                data.get("_caller_user_id"),
                "create_user",
                "user",
                uid,
                "role=" + role_code + "; status=" + status,
                db=txn,
            )

        return uid, raw_pw

    # ============================================================
    # Update
    # ============================================================

    @staticmethod
    def update_user(uid, data, current_user_id=None):
        """Update user info."""
        if "status" in data:
            data["status"] = AdministratorPolicy.normalize_status(data["status"])
        db = BaseService.db()
        existing = UserRepository.find_user_by_id_for_update(uid, db=db)
        if not existing:
            raise NotFoundError("User not found")

        # Position validation
        if "position_id" in data:
            position_id = data["position_id"]
            if position_id:
                if not UserRepository.find_position_by_id(position_id, db=db):
                    raise NotFoundError("Specified position does not exist")
        if "department_id" in data:
            department_id = data["department_id"]
            if department_id and not UserRepository.find_department_by_id(
                department_id, db=db
            ):
                raise NotFoundError("Specified department does not exist")

        # Process validation
        UserService._validate_process_ids(data)

        old_role = UserRepository.get_primary_role_code(uid, db=db) or existing["role"]
        new_role_id, new_role_code = UserService._resolve_update_role(
            uid, data, old_role, current_user_id, db
        )
        if "employee_no" in data:
            data["employee_no"] = (data.get("employee_no") or "").strip()
        sets, params = UserService._build_update_fields(data)

        with BaseService.transaction() as txn:
            existing = UserRepository.find_user_by_id_for_update(uid, db=txn)
            if not existing:
                raise NotFoundError("User not found")
            AdministratorPolicy.protect_admin_accounts(
                current_user_id, [uid], txn
            )
            if (
                uid == current_user_id
                and data.get("status")
                and data["status"] != existing["status"]
            ):
                raise ConflictError("Cannot change your own status")
            if (
                data.get("status") == "inactive"
                and UserRepository.has_admin_assignment(uid, db=txn)
                and UserRepository.count_admin_roles(db=txn) <= 1
            ):
                raise ConflictError("Cannot deactivate the last administrator")
            if "employee_no" in data:
                conflict = UserRepository.find_user_by_employee_no(
                    data["employee_no"], exclude_user_id=uid, db=txn
                )
                if conflict:
                    raise ConflictError("Employee number already exists")
            UserService._sync_user_processes(uid, data, txn)
            UserRepository.update_user_txn(uid, ", ".join(sets), params, db=txn)
            UserService._apply_role_update(uid, old_role, new_role_id, new_role_code, txn)
            updated = PerformanceAssignmentRepository.user_snapshot(uid, db=txn)
            PerformanceAssignmentService.record_user_change(
                existing, updated, created_by=current_user_id, db=txn
            )
            UserService._audit_user_update(
                uid, data, existing, current_user_id, txn
            )

        return True

    # ============================================================
    # Soft Delete / Restore / Permanent Delete
    # ============================================================

    @staticmethod
    def restore_user(uid, actor_id=None):
        with BaseService.transaction() as txn:
            user = UserRepository.find_user_by_id_for_update(uid, db=txn)
            if not user or user["status"] != "deleted":
                raise NotFoundError("User not found or not deleted")
            AdministratorPolicy.protect_admin_accounts(actor_id, [uid], txn)
            if user["purged_at"]:
                raise ConflictError("An anonymized employee identity cannot be restored")
            UserRepository.restore_user_txn(uid, db=txn)
            updated = PerformanceAssignmentRepository.user_snapshot(uid, db=txn)
            PerformanceAssignmentService.record_user_change(user, updated, db=txn)
            UserRepository.insert_audit_log_txn(
                actor_id, "restore_user", "user", uid, "status=active", db=txn
            )
        return True

    @staticmethod
    def permanent_delete_user(uid, actor_id, reason):
        reason = (reason or "").strip()
        if len(reason) < 4:
            raise ValueError("Purge reason must be at least 4 characters")
        with BaseService.transaction() as txn:
            UserService._require_role_administrator(actor_id, txn)
            user = UserRepository.find_deleted_user(uid, db=txn)
            if not user:
                raise ConflictError("Can only permanently delete trashed users")
            if user["purged_at"]:
                raise ConflictError("User identity has already been purged")
            password_hash = bcrypt.hashpw(
                secrets.token_urlsafe(32).encode(), bcrypt.gensalt(rounds=12)
            ).decode()
            UserRepository.anonymize_deleted_user_txn(
                uid, actor_id, reason, password_hash, db=txn
            )
            UserRepository.insert_audit_log_txn(
                actor_id,
                "anonymize_user_identity",
                "user",
                uid,
                "reason=" + reason,
                db=txn,
            )
        return True

    @staticmethod
    def delete_user(uid, current_user_id):
        """Soft-delete user by setting status='deleted'."""
        if uid == current_user_id:
            raise ConflictError("Cannot delete self")

        with BaseService.transaction() as txn:
            before = UserRepository.find_user_by_id_for_update(uid, db=txn)
            if not before:
                raise NotFoundError("User not found")
            AdministratorPolicy.protect_admin_accounts(
                current_user_id, [uid], txn
            )
            if UserRepository.check_admin_role(uid, db=txn):
                if UserRepository.count_admin_roles(db=txn) <= 1:
                    raise ConflictError("Cannot delete the last administrator")
            UserRepository.soft_delete_user_txn(uid, db=txn)
            after = PerformanceAssignmentRepository.user_snapshot(uid, db=txn)
            PerformanceAssignmentService.record_user_change(
                before, after, created_by=current_user_id, db=txn
            )
            UserRepository.insert_audit_log_txn(
                current_user_id,
                "delete_user",
                "user",
                uid,
                "status=deleted",
                db=txn,
            )
        return True

    @staticmethod
    def batch_update_status(ids, status, current_user_id=None):
        """Batch update user status (active/inactive)."""
        if not ids:
            return 0
        ids = UserService._normalize_ids(ids, "ids")
        if current_user_id and current_user_id in ids:
            raise ConflictError("Cannot change own status")
        if status not in ("active", "inactive"):
            raise ValueError("Invalid status")
        with BaseService.transaction() as txn:
            AdministratorPolicy.protect_admin_accounts(
                current_user_id, ids, txn
            )
            if status == "inactive":
                admin_count = UserRepository.count_admin_roles_in_ids(ids, db=txn)
                if (
                    admin_count > 0
                    and UserRepository.count_admin_roles(db=txn) <= admin_count
                ):
                    raise ConflictError("Cannot deactivate all administrators")
            before_rows = {
                row["id"]: row
                for row in UserRepository.find_users_by_ids_for_update(ids, db=txn)
            }
            count = UserRepository.batch_update_status_txn(ids, status, db=txn)
            after_rows = {
                row["id"]: row
                for row in UserRepository.find_users_by_ids_for_update(ids, db=txn)
            }
            for user_id, before in before_rows.items():
                after = after_rows.get(user_id)
                if after and before["status"] != after["status"]:
                    snapshot = PerformanceAssignmentRepository.user_snapshot(
                        user_id, db=txn
                    )
                    PerformanceAssignmentService.record_user_change(
                        before,
                        snapshot,
                        created_by=current_user_id,
                        db=txn,
                    )
            UserRepository.insert_audit_log_txn(
                current_user_id,
                "batch_update_status",
                "user",
                0,
                "status=" + status + "; user_ids=" + ",".join(str(item) for item in ids),
                db=txn,
            )
        return count

    @staticmethod
    def batch_delete_users(ids, current_user_id):
        """Soft-delete multiple users."""
        if not ids:
            return 0
        ids = UserService._normalize_ids(ids, "ids")
        if current_user_id in ids:
            raise ConflictError("Cannot delete self")
        with BaseService.transaction() as txn:
            AdministratorPolicy.protect_admin_accounts(
                current_user_id, ids, txn
            )
            admin_count = UserRepository.count_admin_roles_in_ids(ids, db=txn)
            if (
                admin_count > 0
                and UserRepository.count_admin_roles(db=txn) <= admin_count
            ):
                raise ConflictError("Cannot remove all administrators")
            before_rows = {
                row["id"]: row
                for row in UserRepository.find_users_by_ids_for_update(ids, db=txn)
            }
            count = UserRepository.batch_soft_delete_users_txn(ids, db=txn)
            for user_id, before in before_rows.items():
                snapshot = PerformanceAssignmentRepository.user_snapshot(user_id, db=txn)
                if snapshot and before["status"] != snapshot["status"]:
                    PerformanceAssignmentService.record_user_change(
                        before,
                        snapshot,
                        created_by=current_user_id,
                        db=txn,
                    )
            UserRepository.insert_audit_log_txn(
                current_user_id,
                "batch_delete_users",
                "user",
                0,
                "user_ids=" + ",".join(str(item) for item in ids),
                db=txn,
            )
        return count

    # ============================================================
    # Password Reset & Unlock
    # ============================================================

    @staticmethod
    def reset_password(uid, password=None, actor_id=None):
        """Reset user password with validation and account unlock."""
        new_pw = password if password else secrets.token_urlsafe(8)
        if password:
            if len(password) < 8:
                raise ValueError("Password must be at least 8 characters")
            if not any(c.isalpha() for c in password):
                raise ValueError("Password must contain at least one letter")
            if not any(c.isdigit() for c in password):
                raise ValueError("Password must contain at least one digit")
        hashed = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt(rounds=12)).decode()

        db = BaseService.db()
        existing = UserRepository.find_user_status(uid, db=db)
        if not existing:
            raise NotFoundError("User not found")
        if existing["status"] != "active":
            raise ConflictError("User is disabled, cannot reset password")

        with BaseService.transaction() as txn:
            AdministratorPolicy.require_actual_admin(actor_id, txn)
            UserRepository.reset_password_txn(uid, hashed, db=txn)
            UserRepository.insert_audit_log_txn(
                actor_id,
                "reset_password",
                "user",
                uid,
                "password_changed=true",
                db=txn,
            )

        return new_pw

    @staticmethod
    def unlock_user(uid, actor_id=None):
        """Unlock user (clear brute-force lockout)."""
        db = BaseService.db()
        row = UserRepository.find_user_by_id_basic(uid, db=db)
        if not row:
            raise NotFoundError("User not found")
        with BaseService.transaction() as txn:
            AdministratorPolicy.require_actual_admin(actor_id, txn)
            UserRepository.unlock_user_txn(uid, db=txn)
            UserRepository.insert_audit_log_txn(
                actor_id, "unlock_user", "user", uid, "account_unlocked=true", db=txn
            )
        return row["username"]

    # ============================================================
    # Batch Import
    # ============================================================

    @staticmethod
    def _open_user_import_workbook(filepath):
        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        return wb, wb.active

    @staticmethod
    def _map_user_import_columns(headers):
        col_map = {}
        for index, header in enumerate(headers):
            if header is None:
                continue
            header_text = str(header).strip()
            if header_text in USER_IMPORT_FIELD_MAP:
                col_map[index] = USER_IMPORT_FIELD_MAP[header_text]
            elif header_text.lower() in USER_IMPORT_EN_MAP:
                col_map[index] = USER_IMPORT_EN_MAP[header_text.lower()]
        if not col_map:
            raise ValueError("No valid column headers found")
        return col_map

    @staticmethod
    def _user_import_row_data(row, col_map):
        row_data = {}
        for col_idx, field in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            row_data[field] = str(val).strip() if val is not None else ""
        return row_data

    @staticmethod
    def _normalize_user_import_identity(row_idx, row_data):
        username = row_data.get("username", "")
        name = row_data.get("name", "")
        if not username and not name:
            return None, None, "Row " + str(row_idx) + ": empty username and name"
        return username or name, name or username, None

    @staticmethod
    def _resolve_user_import_position(row_data, pos_map):
        pos_name = row_data.get("position_name", "")
        return pos_map.get(pos_name) if pos_name else None

    @staticmethod
    def _resolve_user_import_role(row_data, caller_id, db):
        role = (row_data.get("role") or "worker").strip()
        role_row = UserRepository.find_active_role_by_code(role, db=db)
        if not role_row:
            raise ValueError("Unknown or inactive role: " + role)
        if role != "worker" and not UserRepository.check_admin_role(caller_id, db=db):
            raise PermissionError("Only administrators can import privileged roles")
        return role, role_row["id"]

    @staticmethod
    def _next_import_employee_no(row_data, db):
        return UserService._next_employee_no(row_data, db)

    @staticmethod
    def _insert_import_user(row_data, username, name, position_id, role, role_id, caller_id, db):
        _, pw_hash = UserService._hash_or_generate_password(row_data)
        uid = UserRepository.insert_user_import_txn(
            username=username, pw_hash=pw_hash, name=name,
            nickname=row_data.get("nickname", ""),
            email=row_data.get("email", ""),
            role="admin" if role == "admin" else "worker",
            employee_no=row_data.get("employee_no", ""),
            phone=row_data.get("phone", ""),
            position_id=position_id, db=db
        )
        UserRepository.insert_user_role_txn(uid, role_id, db=db)
        PerformanceAssignmentService.record_initial_assignment(
            uid, created_by=caller_id, source_type="user_imported", db=db
        )
        return uid

    @staticmethod
    def _import_user_row(row_idx, row, col_map, pos_map, caller_id, db):
        row_data = UserService._user_import_row_data(row, col_map)
        username, name, identity_error = UserService._normalize_user_import_identity(row_idx, row_data)
        if identity_error:
            return False, identity_error
        UserService._validate_username(username, name)
        if UserRepository.find_user_by_username(username, db=db):
            return False, "Row " + str(row_idx) + ": username " + username + " already exists"

        position_id = UserService._resolve_user_import_position(row_data, pos_map)
        if row_data.get("position_name") and not position_id:
            return False, "Row " + str(row_idx) + ": unknown position " + row_data["position_name"]
        role, role_id = UserService._resolve_user_import_role(row_data, caller_id, db)
        row_data["employee_no"] = UserService._next_import_employee_no(row_data, db)
        UserService._insert_import_user(
            row_data, username, name, position_id, role, role_id, caller_id, db
        )
        return True, None

    @staticmethod
    def import_users(filepath, caller_id=None):
        """Import users from .xlsx file. Returns {success, skipped, errors}."""
        wb, ws = UserService._open_user_import_workbook(filepath)
        try:
            headers = [cell.value for cell in ws[1]]
            col_map = UserService._map_user_import_columns(headers)
            db = BaseService.db()
            pos_rows = UserRepository.get_active_positions(db=db)
            pos_map = {r["name"]: r["id"] for r in pos_rows}

            success = 0
            skipped = 0
            errors = []
            with BaseService.transaction() as txn:
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        created, error = UserService._import_user_row(
                            row_idx, row, col_map, pos_map, caller_id, txn
                        )
                        if created:
                            success += 1
                        else:
                            skipped += 1
                            errors.append(error)
                    except PermissionError:
                        raise
                    except Exception as e:
                        skipped += 1
                        errors.append("Row " + str(row_idx) + ": " + str(e)[:80])
                UserRepository.insert_audit_log_txn(
                    caller_id,
                    "import_users",
                    "user",
                    0,
                    "created=" + str(success) + "; skipped=" + str(skipped),
                    db=txn,
                )
        finally:
            wb.close()

        return {
            "success": success,
            "skipped": skipped,
            "total": success + skipped,
            "errors": errors[:20],
            "error_summary": "; ".join(errors[:5]) if errors else "",
        }

    # ============================================================
    # Helpers
    # ============================================================

    @staticmethod
    def get_user(uid):
        """Get single user (without password)."""
        row = UserRepository.find_user_by_id_full(uid)
        if not row:
            raise NotFoundError("User not found")
        u = dict(row)
        u.pop("password", None)
        u.pop("token", None)
        return u

    @staticmethod
    def export_users(role_filter="", keyword="", status=""):
        data = UserService.list_users(1, 9999, role_filter=role_filter, keyword=keyword, status=status)
        users = data.get("users", [])
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["用户名", "姓名", "员工编号", "电话", "邮箱", "岗位", "角色", "状态"])
        for user in users:
            ws.append([
                user.get("username", ""),
                user.get("name", ""),
                user.get("employee_no", ""),
                user.get("phone", ""),
                user.get("email", ""),
                user.get("position_name", ""),
                "管理员" if user.get("role") == "admin" else "员工",
                "正常" if user.get("status") == "active" else "已删除" if user.get("status") == "deleted" else "禁用",
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def get_user_detail(uid):
        user = UserService.get_user(uid)
        db = BaseService.db()
        user["role_names"] = [row["name"] for row in UserRepository.get_user_role_names(uid, db=db)]
        user["assigned_processes"] = [dict(row) for row in UserRepository.get_user_assigned_processes(uid, db=db)]
        stats = UserRepository.get_user_work_stats(uid, db=db)
        user["work_stats"] = dict(stats) if stats else {}
        return user

    @staticmethod
    def list_user_documents(uid):
        return [dict(row) for row in UserRepository.list_user_documents(uid)]

    @staticmethod
    def _employee_document_name(filename):
        filename = str(filename or "").replace("\\", "/").rsplit("/", 1)[-1].strip()
        if not filename or len(filename) > 255 or "\x00" in filename:
            raise ValidationError("附件文件名无效")
        extension = os.path.splitext(filename)[1].lower()
        if extension not in EMPLOYEE_DOCUMENT_ALLOWED_EXTENSIONS:
            raise ValidationError("附件类型不支持")
        return filename, extension

    @staticmethod
    def _validate_employee_document_content(filepath, extension, header):
        if extension == ".pdf" and not header.startswith(b"%PDF-"):
            raise ValidationError("PDF 文件内容与扩展名不一致")
        if extension in {".jpg", ".jpeg"} and not header.startswith(b"\xff\xd8\xff"):
            raise ValidationError("JPEG 文件内容与扩展名不一致")
        if extension == ".png" and not header.startswith(b"\x89PNG\r\n\x1a\n"):
            raise ValidationError("PNG 文件内容与扩展名不一致")
        if extension in {".doc", ".xls"} and not header.startswith(
            b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        ):
            raise ValidationError("Office 文件内容与扩展名不一致")
        if extension in {".docx", ".xlsx"}:
            try:
                with zipfile.ZipFile(filepath) as archive:
                    names = archive.namelist()
            except (OSError, zipfile.BadZipFile) as exc:
                raise ValidationError("Office 文件内容与扩展名不一致") from exc
            required_prefix = "word/" if extension == ".docx" else "xl/"
            if not any(name.startswith(required_prefix) for name in names):
                raise ValidationError("Office 文件内容与扩展名不一致")

    @staticmethod
    def _employee_document_path(upload_dir, stored_name):
        base = os.path.realpath(upload_dir)
        filepath = os.path.realpath(os.path.join(base, stored_name))
        if os.path.commonpath([base, filepath]) != base:
            raise ConflictError("附件存储路径无效")
        return filepath

    @staticmethod
    def _existing_employee_document_path(
        upload_dir, stored_name, legacy_upload_dir=None
    ):
        filepath = UserService._employee_document_path(upload_dir, stored_name)
        if os.path.exists(filepath) or not legacy_upload_dir:
            return filepath
        legacy_path = UserService._employee_document_path(
            legacy_upload_dir, stored_name
        )
        return legacy_path if os.path.exists(legacy_path) else filepath

    @staticmethod
    def upload_user_document(uid, file_storage, doc_type, uploaded_by, upload_dir):
        original_name, extension = UserService._employee_document_name(
            file_storage.filename
        )
        os.makedirs(upload_dir, mode=0o700, exist_ok=True)
        stored_name = str(uid) + "_" + uuid.uuid4().hex + extension
        filepath = UserService._employee_document_path(upload_dir, stored_name)
        descriptor, temp_path = tempfile.mkstemp(prefix=".employee-upload-", dir=upload_dir)
        file_size = 0
        header = b""
        installed = False
        try:
            with os.fdopen(descriptor, "wb") as target:
                while True:
                    chunk = file_storage.stream.read(64 * 1024)
                    if not chunk:
                        break
                    if not header:
                        header = chunk[:16]
                    file_size += len(chunk)
                    if file_size > EMPLOYEE_DOCUMENT_MAX_BYTES:
                        raise ValidationError("员工附件最大允许 20MB")
                    target.write(chunk)
                target.flush()
                os.fsync(target.fileno())
            if file_size == 0:
                raise ValidationError("附件内容不能为空")
            UserService._validate_employee_document_content(
                temp_path, extension, header
            )
            with BaseService.transaction() as txn:
                if not UserRepository.find_user_by_id_basic(uid, db=txn):
                    raise NotFoundError("User not found")
                document_id = UserRepository.insert_user_document_txn(
                    uid,
                    original_name,
                    str(doc_type or "")[:64],
                    stored_name,
                    file_size,
                    uploaded_by,
                    db=txn,
                )
                os.replace(temp_path, filepath)
                installed = True
                os.chmod(filepath, 0o600)
                UserRepository.insert_audit_log_txn(
                    uploaded_by,
                    "upload_document",
                    "user",
                    uid,
                    "document_id=" + str(document_id) + "; size=" + str(file_size),
                    db=txn,
                )
        except Exception:
            if installed and os.path.exists(filepath):
                os.remove(filepath)
            raise
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        return {"filename": original_name, "size": file_size}

    @staticmethod
    def get_user_document_file(uid, doc_id, upload_dir, legacy_upload_dir=None):
        doc = UserRepository.find_user_document(uid, doc_id)
        if not doc:
            raise NotFoundError("Document not found")
        doc = dict(doc)
        filepath = UserService._existing_employee_document_path(
            upload_dir, doc["file_path"], legacy_upload_dir
        )
        if not os.path.exists(filepath):
            raise FileNotFoundError("File not found on disk")
        return doc, filepath

    @staticmethod
    def delete_user_document(
        uid, doc_id, upload_dir, actor_id=None, legacy_upload_dir=None
    ):
        os.makedirs(upload_dir, mode=0o700, exist_ok=True)
        quarantine_dir = os.path.join(upload_dir, ".quarantine")
        os.makedirs(quarantine_dir, mode=0o700, exist_ok=True)
        quarantine_path = None
        filepath = None
        try:
            with BaseService.transaction() as txn:
                row = UserRepository.find_user_document(uid, doc_id, db=txn)
                if not row:
                    raise NotFoundError("Document not found")
                doc = dict(row)
                filepath = UserService._existing_employee_document_path(
                    upload_dir, doc["file_path"], legacy_upload_dir
                )
                if os.path.exists(filepath):
                    quarantine_path = os.path.join(
                        quarantine_dir, uuid.uuid4().hex + ".deleted"
                    )
                    os.replace(filepath, quarantine_path)
                UserRepository.delete_user_document_txn(doc_id, db=txn)
                UserRepository.insert_audit_log_txn(
                    actor_id,
                    "delete_document",
                    "user",
                    uid,
                    "document_id=" + str(doc_id),
                    db=txn,
                )
        except Exception:
            if quarantine_path and os.path.exists(quarantine_path) and filepath:
                os.replace(quarantine_path, filepath)
            raise
        if quarantine_path and os.path.exists(quarantine_path):
            os.remove(quarantine_path)
        return doc
