"""
qr-system - CSV/Excel Bulk Import Routes
Supports: orders, products, customers import
"""
import csv, io, json, re
from datetime import datetime
from flask import request, jsonify, g
from werkzeug.utils import secure_filename

from modules.db import get_db
from modules.app import app
from modules.services.imports_service import ImportsService
from modules.middleware.audit import audit_log
from modules.middleware.auth import check_auth, check_permission, has_permission
from modules.middleware.helpers import get_json_body

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

MAX_PREVIEW_ROWS = 500
MAX_IMPORT_ROWS = 5000
ALLOWED_EXTENSIONS = {'.csv', '.xlsx', '.xls'}

# Field definitions for each import type
IMPORT_FIELDS = {
    'orders': {
        'required': ['order_no'],
        'optional': ['customer', 'product_name', 'quantity', 'plan_start', 'plan_end', 'deadline', 'remark', 'product_code'],
        'label': '订单',
        'perms': 'orders:create',
    },
    'products': {
        'required': ['product_name'],
        'optional': ['model', 'spec', 'upper_opening', 'plate_thickness', 'style', 'description', 'unit'],
        'label': '产品',
        'perms': 'products:create',
    },
    'customers': {
        'required': ['name'],
        'optional': ['contact', 'phone', 'address', 'email'],
        'label': '客户',
        'perms': 'customers:create',
    },
}


def _parse_file(file_storage):
    """Parse uploaded CSV or Excel file, return list of dicts."""
    filename = (file_storage.filename or '').lower()
    rows = []

    if filename.endswith('.csv'):
        text = file_storage.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            cleaned = {k.strip(): v.strip() if v else '' for k, v in row.items() if k and k.strip()}
            if any(cleaned.values()):
                rows.append(cleaned)
    elif filename.endswith(('.xlsx', '.xls')):
        if not HAS_OPENPYXL:
            raise ValueError('Server missing openpyxl library; Excel not supported')
        wb = openpyxl.load_workbook(io.BytesIO(file_storage.read()), read_only=True)
        ws = wb.active
        headers = []
        for i, cell in enumerate(ws[1]):
            h = str(cell.value).strip() if cell.value else ''
            headers.append(h)
        if not headers:
            wb.close()
            return []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for j, val in enumerate(row):
                if j < len(headers) and headers[j]:
                    d[headers[j]] = str(val).strip() if val is not None else ''
            if any(d.values()):
                rows.append(d)
        wb.close()
    else:
        raise ValueError('Unsupported file format; use .csv, .xlsx, or .xls')

    return rows


@app.route('/api/import/preview', methods=['POST'])
@check_auth
@check_permission("orders:create")
def import_preview():
    """Upload file and return preview of parsed rows."""
    import_type = request.form.get('type', 'orders')
    if import_type not in IMPORT_FIELDS:
        return jsonify({'error': 'Invalid import type'}), 400

    perm = IMPORT_FIELDS[import_type]['perms']
    if not has_permission(g.current_user, perm):
        return jsonify({'error': 'No permission'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty file'}), 400

    safe_name = secure_filename(file.filename)
    ext = '.' + safe_name.rsplit('.', 1)[-1].lower() if '.' in safe_name else ''
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'error': f'Unsupported format: {ext}. Use .csv, .xlsx'}), 400

    try:
        rows = _parse_file(file)
    except Exception as e:
        return jsonify({'error': f'Parse error: {str(e)}'}), 400

    if not rows:
        return jsonify({'error': 'No data rows found'}), 400

    total = len(rows)
    preview = rows[:MAX_PREVIEW_ROWS]

    # Validate headers
    field_info = IMPORT_FIELDS[import_type]
    required = set(field_info['required'])
    optional = set(field_info['optional'])
    all_valid = required | optional

    if preview:
        headers = list(preview[0].keys())
        unknown = [h for h in headers if h not in all_valid]
        missing = [h for h in required if h not in headers]
    else:
        unknown = []
        missing = list(required)

    import re
    # Count valid rows
    valid_count = 0
    errors = []
    for i, row in enumerate(preview):
        row_ok = True
        for req in required:
            if not row.get(req, '').strip():
                errors.append({'row': i + 2, 'field': req, 'error': 'Required field empty'})
                row_ok = False
        if row_ok:
            valid_count += 1

    return jsonify({
        'type': import_type,
        'total_rows': total,
        'preview_rows': len(preview),
        'valid_rows': valid_count,
        'headers': list(preview[0].keys()) if preview else [],
        'missing_required': missing,
        'unknown_headers': unknown,
        'errors': errors[:50],
        'preview': preview[:20],
    })


@app.route('/api/import/orders', methods=['POST'])
@check_auth
@check_permission('orders:create')
def bulk_import_orders():
    """Import orders from preview-confirmed data."""
    data = get_json_body()
    rows = data.get('rows', [])
    if not rows:
        return jsonify({'error': 'No data'}), 400
    if len(rows) > MAX_IMPORT_ROWS:
        return jsonify({'error': f'Max {MAX_IMPORT_ROWS} rows per import'}), 400

    db = get_db()
    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows):
        order_no = (row.get('order_no') or '').strip()
        if not order_no:
            errors.append({'row': i + 1, 'error': 'Missing order_no'})
            skipped += 1
            continue

        # Check duplicate
        existing = ImportsService.check_order_exists(order_no)
        if existing:
            errors.append({'row': i + 1, 'order_no': order_no, 'error': 'Duplicate order_no'})
            skipped += 1
            continue

        try:
            ImportsService.insert_order({
                'order_no': order_no, 'product_name': product_name, 'product_code': product_code,
                'customer': customer, 'quantity': quantity, 'status': status,
                'plan_start': plan_start, 'plan_end': plan_end, 'deadline': deadline, 'remark': remark
            })
            imported += 1
        except Exception as e:
            errors.append({'row': i + 1, 'order_no': order_no, 'error': str(e)[:100]})
            skipped += 1

    audit_log('import_orders', detail=f'imported={imported} skipped={skipped}')
    return jsonify({
        'imported': imported,
        'skipped': skipped,
        'errors': errors[:50],
    })


@app.route('/api/import/products', methods=['POST'])
@check_auth
@check_permission('products:create')
def bulk_import_products():
    """Import products from preview-confirmed data."""
    data = get_json_body()
    rows = data.get('rows', [])
    if not rows:
        return jsonify({'error': 'No data'}), 400
    if len(rows) > MAX_IMPORT_ROWS:
        return jsonify({'error': f'Max {MAX_IMPORT_ROWS} rows per import'}), 400

    db = get_db()
    imported = 0
    errors = []

    for i, row in enumerate(rows):
        product_name = (row.get('product_name') or '').strip()
        if not product_name:
            errors.append({'row': i + 1, 'error': 'Missing product_name'})
            continue

        try:
            ImportsService.insert_product({
                'product_name': product_name, 'model': model, 'spec': spec,
                'category': category, 'weight': weight, 'price': price
            })
            imported += 1
        except Exception as e:
            errors.append({'row': i + 1, 'error': str(e)[:100]})
            skipped += 1

    audit_log('import_products', detail=f'imported={imported}')
    return jsonify({
        'imported': imported,
        'errors': errors[:50],
    })


@app.route('/api/import/customers', methods=['POST'])
@check_auth
@check_permission('customers:create')
def bulk_import_customers():
    """Import customers from preview-confirmed data."""
    data = get_json_body()
    rows = data.get('rows', [])
    if not rows:
        return jsonify({'error': 'No data'}), 400
    if len(rows) > MAX_IMPORT_ROWS:
        return jsonify({'error': f'Max {MAX_IMPORT_ROWS} rows per import'}), 400

    db = get_db()
    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows):
        name = (row.get('name') or '').strip()
        if not name:
            errors.append({'row': i + 1, 'error': 'Missing name'})
            skipped += 1
            continue

        existing = ImportsService.check_customer_exists(name)
        if existing:
            errors.append({'row': i + 1, 'name': name, 'error': 'Duplicate name'})
            skipped += 1
            continue

        try:
            ImportsService.insert_customer({
                'name': name, 'contact': contact, 'phone': phone,
                'address': address, 'remark': remark
            })
            imported += 1
        except Exception as e:
            errors.append({'row': i + 1, 'name': name, 'error': str(e)[:100]})
            skipped += 1

    audit_log('import_customers', detail=f'imported={imported} skipped={skipped}')
    return jsonify({
        'imported': imported,
        'skipped': skipped,
        'errors': errors[:50],
    })


# Download template endpoint
@app.route('/api/import/template/<import_type>', methods=['GET'])
@check_auth
def import_template(import_type):
    """Download a blank CSV template for the given type."""
    if import_type not in IMPORT_FIELDS:
        return jsonify({'error': 'Invalid type'}), 400

    field_info = IMPORT_FIELDS[import_type]
    headers = field_info['required'] + field_info['optional']

    import csv, io as io_mod
    output = io_mod.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    # Add one sample row
    samples = {
        'orders': ['ORD-2026-001', 'ABC Company', 'Widget', '100', '2026-01-01', '2026-01-15', '2026-01-20', 'Sample remark', 'WIDGET-01'],
        'products': ['Widget', 'W-100', 'Standard', '330', '25', 'Standard', 'Sample product', '个'],
        'customers': ['ABC Company', 'John Doe', '13800138000', 'Shanghai', 'john@abc.com'],
    }
    sample = samples.get(import_type, [''] * len(headers))
    # Only output sample values for fields that exist
    writer.writerow(sample[:len(headers)])

    from flask import Response
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={import_type}_template.csv'}
    )