"""qr-system - Excel Export Utility"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, headers, row=1):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

def auto_width(ws, min_width=10, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted

def export_orders_to_excel(orders, filename=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = ["ID", "Order No", "Customer", "Product", "Quantity",
               "Status", "Plan Start", "Plan End", "Deadline", "Created"]
    style_header(ws, headers)

    for row_idx, order in enumerate(orders, 2):
        values = [
            order.get("id", ""),
            order.get("order_no", ""),
            order.get("customer", ""),
            order.get("product_name", ""),
            order.get("quantity", 0),
            order.get("status", ""),
            order.get("plan_start", ""),
            order.get("plan_end", ""),
            order.get("deadline", ""),
            order.get("created_at", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

    auto_width(ws)

    if filename:
        wb.save(filename)
        return filename
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_work_records_to_excel(records, filename=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Work Records"

    headers = ["ID", "Order No", "Process", "Worker", "Serial No",
               "Quantity", "Type", "Remark", "Reported At"]
    style_header(ws, headers)

    for row_idx, rec in enumerate(records, 2):
        values = [
            rec.get("id", ""),
            rec.get("order_no", ""),
            rec.get("process_name", ""),
            rec.get("worker_name", ""),
            rec.get("serial_no", ""),
            rec.get("quantity", 0),
            rec.get("type", ""),
            rec.get("remark", ""),
            rec.get("created_at", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

    auto_width(ws)

    if filename:
        wb.save(filename)
        return filename
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_inventory_to_excel(items, filename=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["ID", "Product Code", "Product Name", "Serial No",
               "Status", "Location", "Updated"]
    style_header(ws, headers)

    for row_idx, item in enumerate(items, 2):
        values = [
            item.get("id", ""),
            item.get("product_code", ""),
            item.get("product_name", ""),
            item.get("serial_no", ""),
            item.get("status", ""),
            item.get("location", ""),
            item.get("updated_at", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

    auto_width(ws)

    if filename:
        wb.save(filename)
        return filename
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output